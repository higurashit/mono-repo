import polars as pl
import boto3
from openpyxl import Workbook, load_workbook
from processor import prepare_output_df

s3 = boto3.client('s3', region_name='ap-northeast-1')

def output_excel(bucket_name, object_path, local_directory_path, dst):
    df = dst["df"]
    file_path = f'{local_directory_path}{dst["output"]["path"]}'
    s3_file_path = f'{object_path}{dst["output"]["path"]}'
    
    # 元データをExcelに貼り付け
    title = f'元データ_{dst["data_name"]}'
    try:
      wb = load_workbook(file_path)
    except FileNotFoundError:
      wb = Workbook()
      ws = wb.active
      del wb[ws.title]

    # 元データを出力
    ws = ws_append(title, wb, df)

    # 出力シート情報に従って列順を変換して出力
    definition:pl.DataFrame = dst["definition"]
    columuns = definition.columns
    sheets = definition.select(columuns[9:18])
    # print(f'sheets: {sheets}')
    for sheet_name, sheet_row_nums in sheets.to_dict().items():
      # 出力用に列の順序や位置を変更
      col_names = [f["name"] for f in dst["field_def"]]
      ret_df = prepare_output_df(sheet_row_nums, col_names, col_names, dst)
      # 出力シート情報が空の場合はスキップ
      if len(ret_df) == 0:
        continue
      # 指定のExcelシートにクエリ結果を出力
      ws = ws_append(sheet_name, wb, ret_df)

    try:
      # Excelを保存
      wb.save(file_path)
      print(f'Excelファイルを保存しました: {file_path}')
      s3.upload_file(file_path, bucket_name, s3_file_path)
      print(f'ExcelファイルをS3に保存しました: s3://{bucket_name}/{s3_file_path}')
    except PermissionError as pe:
      print(f'PermissionError: {file_path} に書き込みできません。ファイルを削除するか閉じてください。')
      raise pe

def ws_append(sheet_name, wb:Workbook, df:pl.DataFrame):
    # 同名のシートが存在する場合は削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]  # シートを削除
    # 指定のExcelシートにクエリ結果を出力
    ws = wb.create_sheet(title=sheet_name)
    ws.append(df.columns)
    for d in df.to_dicts():
      ws.append(list(d.values()))
    
    return ws

def output_csv(bucket_name, object_path, local_directory_path, dst):
    file_path = f'{local_directory_path}{dst["output"]["path"]}'
    s3_file_path = f'{object_path}{dst["output"]["path"]}'
    definition:pl.DataFrame = dst["definition"]
    columns = definition.columns
    # CSVは一番左の出力列（10列目）を採用
    _, row_nums = definition.select(columns[9]).to_dict().popitem()
    col_names = [f["name"] for f in dst["field_def"]]
    # CSVはエイリアスを物理名で指定
    aliases = [f["physical_name"] for f in dst["field_def"]]

    # print(f'definition: {definition}, columns: {columns}, row_nums: {row_nums}')
    # CSVはエイリアスを物理名で指定
    ret_df:pl.DataFrame = prepare_output_df(row_nums, col_names, aliases, dst)

    try:
      # 定義に従って列順を変換して出力
      # CSVに出力して保存
      ret_df.write_csv(file=file_path, separator=',')
      print(f'CSVファイルを保存しました: {file_path}')
      s3.upload_file(file_path, bucket_name, s3_file_path)
      print(f'CSVファイルをS3に保存しました: s3://{bucket_name}{s3_file_path}')
    except PermissionError as pe:
      print(f'PermissionError: {file_path} に書き込みできません。ファイルを削除するか閉じてください。')
      raise pe