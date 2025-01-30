import polars as pl
from openpyxl import Workbook, load_workbook
from processor import prepare_output_df

def output_excel(dst):
    df = dst["df"]
    file_path = dst["output"]["path"]
    
    # 元データをExcelに貼り付け
    title = f'元データ_{dst["data_name"]}'
    try:
      wb = load_workbook(file_path)
      ws = wb.create_sheet(title=title)
    except FileNotFoundError:
      wb = Workbook()
      ws = wb.active
      ws.title = title
    
    ws.append(df.columns)
    for d in df.to_dicts():
      ws.append(list(d.values()))

    # 定義に従って列順を変換して出力
    definition = dst["definition"]
    columuns = definition.columns
    sheets = definition.select(columuns[9:18])
    # print(f'sheets: {sheets}')
    for sheet_name, sheet_row_nums in sheets.to_dict().items():
      # 列変換用のクエリを取得
      col_names = [f["name"] for f in dst["field_def"]]
      # 指定のExcelシートにクエリ結果を出力
      ret_df = prepare_output_df(sheet_row_nums, col_names, col_names, dst)
      if len(ret_df) == 0:
        continue
      ws = wb.create_sheet(title=sheet_name)
      ws.append(ret_df.columns)
      for d in ret_df.to_dicts():
        ws.append(list(d.values()))
    # Excelを保存
    wb.save(file_path)

def output_csv(dst):
    file_path = dst["output"]["path"]
    definition = dst["definition"]
    columns = definition.columns
    # CSVは一番左の出力列（10列目）を採用
    _, row_nums = definition.select(columns[9]).to_dict().popitem()
    col_names = [f["name"] for f in dst["field_def"]]
    # CSVはエイリアスを物理名で指定
    aliases = [f["physical_name"] for f in dst["field_def"]]

    # print(f'definition: {definition}, columns: {columns}, row_nums: {row_nums}')
    # CSVはエイリアスを物理名で指定
    ret_df = prepare_output_df(row_nums, col_names, aliases, dst)

    # 定義に従って列順を変換して出力
    # CSVに出力して保存
    file_path = dst["output"]["path"]
    ret_df.write_csv(file=file_path, separator=',')