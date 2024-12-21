import json
import os
import boto3
import openpyxl as xl
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html
from datetime import datetime, timedelta, timezone
print('Loading function')

# define
JST = timezone(timedelta(hours=+9), "JST")
# s3 = boto3.client("s3")
# EXCEL_BUCKET_NAME = os.getenv("EXCEL_BUCKET_NAME")
# EXCEL_KEY_NAME = "sample.xlsx"

def handler(event, context):
    
    try:
        # Extract input parameters
        # Ex. migration_definition_path: s3://siruko-cloudformation-templetes/migrationTest.json
        target_sheet_names = event.get('target_sheet_names', [])
        print('hoge')

        # 移行定義ファイルが存在するか
        # local_excel_path = get_excel(EXCEL_BUCKET_NAME, EXCEL_KEY_NAME)
        local_excel_path = './sample.xlsx'

        # 移行対象の定義を取得
        wsheets = load_excel_data(local_excel_path, target_sheet_names)

        # 各シートのフォーマットチェック
        check_format(wsheets)

        # 移行定義ファイルから移行元データリストを生成
        src_file_list = []

        for ws in wsheets:
            src_file = get_src_file(ws)
            src_file_list.append(src_file)

        print(f'src_file_list: {src_file_list}')
        return {
            'src_file_list': src_file_list
        }

    except Exception as e:
        print(e)
        return {
            'statusCode': 400,
            'body': json.dumps({
                'error': str(e)
            })
        }

def get_excel(bucket, key):
    local_file_path = os.path.join("/tmp", key)

    s3.download_file(bucket, key, local_file_path)

    if os.path.exists(local_file_path):
        return local_file_path

    raise FileNotFoundError(f"File not found at: {local_file_path}")

def load_excel_data(local_file_path, target_sheet_names):
    # target_sheet_names = ['s1', 's4']
    # load excel
    wb = xl.load_workbook(filename=local_file_path)
    print(wb.sheetnames)
    # set target sheets
    wsheets = get_target_sheets(wb, target_sheet_names)

    ret = []
    for ws in wsheets:
        ret.append(ws)
        print(ws.calculate_dimension())
    
    print(f'ret: {ret}')
    return ret

def get_target_sheets(wb, target_sheet_names):
    target_sheets = []
    if len(target_sheet_names) > 0:
        for target_sheet_name in target_sheet_names:
            if target_sheet_name in wb.sheetnames:
                print(f'target_sheet_name: {target_sheet_name}')
                target_sheets.append(wb[target_sheet_name])
    else:
        target_sheets = wb.worksheets

    if len(target_sheets) == 0:
        print('hoge8')
        raise Exception(f"移行対象が存在しません。wb:{wb.worksheets}, target: {target_sheet_names}")
    
    return target_sheets

def check_format(wsheets):
    for ws in wsheets:
        # 元ファイル名が空
        if ws.cell(row=1, column=1).value is None:
            raise Exception(f"元ファイル名が空です。ws:{ws}")
        # 元データに空の項目名、型、桁がある。
        if ws.cell(row=1, column=1).value is None:
            raise Exception(f"元データに空の項目名、型、桁があります。ws:{ws}")

def get_src_file(ws, file_path_row = 1, file_path_column = 1):
    KEY_PREFIX = 'src_files'
    src_file = ws.cell(file_path_row, file_path_column).value
    if src_file is None:
        raise Exception(f"移行元ファイル名が空です。ws:{ws}")

    return f'{KEY_PREFIX}/{src_file}'

if __name__ == "__main__":
    event = {
        # "target_sheet_names": ["Sheet1"]
    }
    handler(event, None)